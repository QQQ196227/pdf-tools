import sys
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def pdf_to_xlsx(input_path, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "PDF内容"

    # 样式
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    with pdfplumber.open(input_path) as pdf:
        row_num = 1
        for page_num, page in enumerate(pdf.pages, 1):
            # 页面标题
            ws.cell(row=row_num, column=1, value=f"--- 第 {page_num} 页 ---").font = Font(bold=True, size=14)
            row_num += 1

            # 提取表格
            tables = page.extract_tables()
            if tables:
                for table_idx, table in enumerate(tables):
                    if table_idx > 0:
                        row_num += 1  # 多个表格间空行
                    for row_data in table:
                        if row_data:
                            for col_num, cell in enumerate(row_data, 1):
                                c = ws.cell(row=row_num, column=col_num, value=str(cell).strip() if cell else "")
                                c.border = thin_border
                                c.alignment = Alignment(wrap_text=True)
                            row_num += 1

            # 提取文本
            text = page.extract_text()
            if text:
                for line in text.split('\n'):
                    if line.strip():
                        ws.cell(row=row_num, column=1, value=line)
                        row_num += 1

            row_num += 1  # 页面间空行

    # 自动调整列宽
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    wb.save(output_path)
    return True

if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("Usage: pdf_to_xlsx.py input.pdf output.xlsx")
        sys.exit(1)
    try:
        pdf_to_xlsx(sys.argv[1], sys.argv[2])
        print("OK")
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)
